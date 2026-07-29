"""ChinaDrugTrials 官方导出文件的读取、比较与缓存更新。"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from bs4 import BeautifulSoup

from .clinical_registry import load_china_drug_trials_cache, normalize_china_drug_trials_record
from .io import atomic_write_json


CHANGE_FIELDS = (
    "title",
    "drug_name",
    "indication",
    "status",
    "status_raw",
    "phase",
    "sponsor",
    "registered_date",
    "official_url",
)

HEADER_ALIASES = {
    "registry_id": (
        "registry_id", "registration_number", "ctr_number",
        "登记号", "试验登记号", "登记编号", "临床试验登记号",
    ),
    "title": (
        "title", "study_title", "试验题目", "临床试验题目", "试验名称",
    ),
    "drug_name": (
        "drug_name", "drug", "intervention_name",
        "药物名称", "试验药物名称", "药品名称", "试验药品",
    ),
    "indication": (
        "indication", "disease", "target_disease",
        "适应症", "适应证", "目标适应症", "目标适应证",
    ),
    "status": (
        "status", "recruitment_status", "study_status",
        "试验状态", "招募状态", "试验进展", "公示状态",
    ),
    "phase": (
        "phase", "study_phase", "试验分期", "临床试验分期", "分期",
    ),
    "sponsor": (
        "sponsor", "applicant", "company",
        "申办者", "申办者名称", "申请人", "申请人名称",
    ),
    "registered_date": (
        "registered_date", "registration_date", "first_public_date",
        "登记日期", "首次公示日期", "首次公示信息日期", "公示日期",
    ),
    "official_url": (
        "official_url", "url", "source_url", "详情链接", "公示链接", "链接",
    ),
}


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _header_key(value: Any) -> str:
    return re.sub(r"[\s_\-—–:：()（）/\\]+", "", _stringify(value)).lower()


ALIAS_LOOKUP = {
    _header_key(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _rows_to_records(rows: Iterable[Iterable[Any]]) -> list[dict[str, str]]:
    materialized = [list(row) for row in rows]
    if not materialized:
        return []
    header_index = -1
    mapped_headers: list[str] = []
    for index, row in enumerate(materialized[:30]):
        candidate = [ALIAS_LOOKUP.get(_header_key(cell), "") for cell in row]
        if len({value for value in candidate if value}) >= 2 and "registry_id" in candidate:
            header_index = index
            mapped_headers = candidate
            break
    if header_index < 0:
        raise ValueError("未识别 ChinaDrugTrials 表头；至少需要登记号和一个试验字段")

    records: list[dict[str, str]] = []
    for row in materialized[header_index + 1:]:
        record = {
            key: _stringify(row[column])
            for column, key in enumerate(mapped_headers)
            if key and column < len(row) and _stringify(row[column])
        }
        if record.get("registry_id"):
            records.append(record)
    return records


def _read_json(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("records") or payload.get("trials") or payload.get("data") or []
    if not isinstance(payload, list):
        raise ValueError("JSON 导出必须包含记录数组")
    return [dict(item) for item in payload if isinstance(item, dict)]


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    return _rows_to_records(rows)


def _read_html_table(path: Path) -> list[dict[str, str]]:
    soup = BeautifulSoup(path.read_bytes(), "html.parser")
    records: list[dict[str, str]] = []
    for table in soup.find_all("table"):
        rows = [
            [_stringify(cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])]
            for row in table.find_all("tr")
        ]
        try:
            records.extend(_rows_to_records(rows))
        except ValueError:
            continue
    if not records:
        raise ValueError("HTML/XLS 导出中没有可识别的试验表格")
    return records


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        try:
            records.extend(_rows_to_records(sheet.iter_rows(values_only=True)))
        except ValueError:
            continue
    if not records:
        raise ValueError("XLSX 中没有可识别的试验表格")
    return records


def _read_xls(path: Path) -> list[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("读取二进制 XLS 需要安装 xlrd") from exc
    workbook = xlrd.open_workbook(path)
    records: list[dict[str, str]] = []
    for sheet in workbook.sheets():
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        try:
            records.extend(_rows_to_records(rows))
        except ValueError:
            continue
    if not records:
        raise ValueError("XLS 中没有可识别的试验表格")
    return records


def load_export_records(path: Path) -> list[dict[str, Any]]:
    """读取官方 JSON/CSV/XLS/XLSX；兼容以 .xls 命名的 HTML 导出。"""
    if not path.is_file():
        raise FileNotFoundError(path)
    prefix = path.read_bytes()[:16].lstrip()
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _read_json(path)
    if suffix in {".csv", ".tsv"}:
        return _read_csv(path)
    if prefix.startswith(b"PK\x03\x04") or suffix == ".xlsx":
        return _read_xlsx(path)
    if prefix.startswith((b"<", b"\xef\xbb\xbf<")):
        return _read_html_table(path)
    if prefix.startswith(b"\xd0\xcf\x11\xe0") or suffix == ".xls":
        return _read_xls(path)
    raise ValueError(f"不支持的 ChinaDrugTrials 文件格式: {path.name}")


def _is_mg_record(record: dict[str, Any]) -> bool:
    text = " ".join(
        str(record.get(key) or "")
        for key in ("title", "indication")
    ).lower()
    return "重症肌无力" in text or "myasthenia gravis" in text


def normalize_export_records(raw_records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """归一化、筛选 MG 并按登记号去重。"""
    by_id: dict[str, dict[str, Any]] = {}
    for raw in raw_records:
        record = normalize_china_drug_trials_record(raw)
        if not record["registry_id"] or not record["title"] or not _is_mg_record(record):
            continue
        if not record.get("official_url"):
            record["official_url"] = (
                "https://www.chinadrugtrials.org.cn/"
                "clinicaltrials.searchlistpage.dhtml?searchtype=keyword&keyword="
                + record["registry_id"]
            )
        by_id[record["registry_id"]] = record
    if not by_id:
        raise ValueError("导出文件中没有有效的重症肌无力试验记录")
    return sorted(by_id.values(), key=lambda item: item["registry_id"], reverse=True)


def compare_records(
    old_records: Iterable[dict[str, Any]],
    new_records: Iterable[dict[str, Any]],
) -> dict[str, Any]:
    """按登记号比较新增、更新、移除和未变化记录。"""
    old_by_id = {str(item.get("registry_id") or ""): item for item in old_records if item.get("registry_id")}
    new_by_id = {str(item.get("registry_id") or ""): item for item in new_records if item.get("registry_id")}
    added_ids = sorted(set(new_by_id) - set(old_by_id), reverse=True)
    removed_ids = sorted(set(old_by_id) - set(new_by_id), reverse=True)
    updated = []
    unchanged_count = 0
    for registry_id in sorted(set(old_by_id) & set(new_by_id), reverse=True):
        changes = {}
        for field in CHANGE_FIELDS:
            before = _stringify(old_by_id[registry_id].get(field))
            after = _stringify(new_by_id[registry_id].get(field))
            if before != after:
                changes[field] = {"before": before, "after": after}
        if changes:
            updated.append({"registry_id": registry_id, "changes": changes})
        else:
            unchanged_count += 1
    return {
        "old_count": len(old_by_id),
        "new_count": len(new_by_id),
        "added_count": len(added_ids),
        "updated_count": len(updated),
        "removed_count": len(removed_ids),
        "unchanged_count": unchanged_count,
        "added": [new_by_id[registry_id] for registry_id in added_ids],
        "updated": updated,
        "removed": [old_by_id[registry_id] for registry_id in removed_ids],
    }


def import_china_drug_trials_exports(
    cache_path: Path,
    input_paths: list[Path],
    *,
    changes_path: Path | None = None,
    allow_large_drop: bool = False,
    dry_run: bool = False,
    now: datetime | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """合并官方导出、生成差异并在安全检查通过后原子更新缓存。"""
    if not input_paths:
        raise ValueError("至少需要一个 ChinaDrugTrials 导出文件")
    cached = load_china_drug_trials_cache(cache_path)
    raw_records = []
    for path in input_paths:
        raw_records.extend(load_export_records(path))
    records = normalize_export_records(raw_records)
    changes = compare_records(cached.get("records") or [], records)
    old_count = changes["old_count"]
    if old_count and len(records) < max(1, int(old_count * 0.6)) and not allow_large_drop:
        raise ValueError(
            f"新文件仅有 {len(records)} 条，低于旧缓存 {old_count} 条的 60%；"
            "请确认提供的是完整导出，或显式使用 --allow-large-drop"
        )

    generated_at = (now or datetime.now(timezone.utc)).astimezone(timezone.utc).isoformat()
    input_meta = [
        {
            "name": path.name,
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        }
        for path in input_paths
    ]
    payload = {
        "schema_version": "1.0",
        "source": "ChinaDrugTrials.org.cn",
        "source_url": "https://www.chinadrugtrials.org.cn/",
        "mode": "manual",
        "generated_at": generated_at,
        "query": "重症肌无力",
        "total": len(records),
        "inputs": input_meta,
        "records": records,
    }
    change_payload = {
        "schema_version": "1.0",
        "source": "ChinaDrugTrials.org.cn",
        "generated_at": generated_at,
        "inputs": input_meta,
        **changes,
    }
    if not dry_run:
        atomic_write_json(cache_path, payload)
        if changes_path:
            atomic_write_json(changes_path, change_payload)
    return payload, change_payload
